import openpyxl as op
import pprint as pp
filename = 'Бланк заказа.xlsx'
subcategory_dict = {}
"""Parsing dates from blank"""
# function from library openpyxl wb2 = load_workbook('test.xls')
wb = op.load_workbook(filename, data_only=True)
# chose a right sheet from exel file
sheet = wb.active
# use all lines and group list with item number, and write to file
# find count of lines
max_rows = sheet.max_row
# try to check our max_rows - print(max_rows)

for i in range(7, max_rows+1):
    sku = sheet.cell(row=i, column=2).value
    subcategory = sheet.cell(row=i, column=12).value
    if not sku:
        continue
    # create dictionaries with all articles (one key - few values)
    if subcategory not in subcategory_dict:
        subcategory_dict[subcategory] = [sku]
    else:
        subcategory_dict[subcategory].append(sku)
# checking our dict via library pprint - for beautiful outputting
pp.pprint(subcategory_dict)

# sorting dictionary by keys
sorteddict = dict(sorted(subcategory_dict.items()))

"""write to file"""
with open('subcategories.ini', 'w', encoding='utf-8') as myfile:
    for key, value in sorteddict.items():
        string_values = ', '.join(value)
        string_to_write = key + ' = ' + string_values + '\n'
        myfile.write(string_to_write)









